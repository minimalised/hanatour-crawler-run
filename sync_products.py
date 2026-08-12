import os
import json
import io
import openpyxl
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# 실행 대상 최상위 폴더 ID
TARGET_FOLDER_ID = "1tTZ8KjqzxV-9t9Vje5jQma8gRmyb8Wl3"

def get_all_target_folder_ids(drive_service, root_id):
    """지정한 폴더 및 그 하위 폴더들의 ID를 모두 수집"""
    folder_ids = [root_id]
    queue = [root_id]
    
    while queue:
        current_id = queue.pop(0)
        query = (
            f"'{current_id}' in parents and "
            "mimeType = 'application/vnd.google-apps.folder' and "
            "trashed = false"
        )
        results = drive_service.files().list(
            q=query,
            corpora='allDrives',
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            fields="files(id)"
        ).execute()
        
        for folder in results.get('files', []):
            folder_ids.append(folder['id'])
            queue.append(folder['id'])
            
    return folder_ids

def sync_sheets():
    # 1. 인증 정보 구성
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.readonly"
    ]
    service_account_info = json.loads(os.environ["GOOGLE_JSON_RAW"])
    creds = Credentials.from_service_account_info(service_account_info, scopes=scopes)
    
    drive_service = build('drive', 'v3', credentials=creds)
    gspread_client = gspread.authorize(creds)

    print(f"[1/4] Scanning target folder tree ({TARGET_FOLDER_ID})...")
    target_folder_ids = get_all_target_folder_ids(drive_service, TARGET_FOLDER_ID)
    print(f"Total target folders found: {len(target_folder_ids)}")

    # 2. 지정된 폴더 범위 내에서만 '대표상품' 키워드가 들어간 최신 .xlsx 파일 조회
    parents_conditions = " or ".join([f"'{f_id}' in parents" for f_id in target_folder_ids])
    query = (
        f"({parents_conditions}) and "
        "name contains '대표상품' and "
        "name contains '.xlsx' and "
        "trashed = false"
    )
    
    results = drive_service.files().list(
        q=query,
        corpora='allDrives',
        includeItemsFromAllDrives=True,
        supportsAllDrives=True,
        orderBy="modifiedTime desc",
        pageSize=5,
        fields="files(id, name, modifiedTime)"
    ).execute()
    
    files = results.get('files', [])

    if not files:
        raise FileNotFoundError(f"지정한 폴더({TARGET_FOLDER_ID}) 내에서 '대표상품' .xlsx 파일을 찾을 수 없습니다.")

    latest_file = files[0]
    print(f"[2/4] Target File Found: '{latest_file['name']}' (ID: {latest_file['id']}, Modified: {latest_file['modifiedTime']})")

    # 3. 파일 메모리 다운로드
    request = drive_service.files().get_media(fileId=latest_file['id'])
    file_stream = io.BytesIO()
    downloader = MediaIoBaseDownload(file_stream, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    
    file_stream.seek(0)

    # 4. openpyxl로 '대표상품리스트' 시트 데이터 읽기
    workbook = openpyxl.load_workbook(file_stream, data_only=True)
    sheet_name = "대표상품리스트"
    
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"엑셀 파일 내 '{sheet_name}' 시트가 없습니다. (존재하는 시트: {workbook.sheetnames})")
    
    excel_sheet = workbook[sheet_name]
    data = []
    for row in excel_sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append([str(cell) if cell is not None else "" for cell in row])

    print(f"[3/4] Successfully extracted {len(data)} rows from '{latest_file['name']}'.")

    # 5. 타겟 스프레드시트에 쓰기
    target_id = os.environ["TARGET_SPREADSHEET_ID"]
    print("[4/4] Writing data to target Google Sheet ('대표상품리스트')...")
    target_sheet = gspread_client.open_by_key(target_id).worksheet("대표상품리스트")
    
    target_sheet.clear()
    target_sheet.update(range_name="A1", values=data)
    print("Done! Sync completed successfully.")

if __name__ == "__main__":
    sync_sheets()
